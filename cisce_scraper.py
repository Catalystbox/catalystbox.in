"""
CatalystBox CISCE School Scraper v2
"""

import requests
import re
import csv
import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"}
REQUEST_DELAY = 1.0
TIMEOUT = 20
SESSION = requests.Session()
SESSION.headers.update(HEADERS)
BASE = "https://www.prokerala.com"

STATE_CODE = {
    "Andaman and Nicobar Islands":"AN","Andhra Pradesh":"AP","Arunachal Pradesh":"AR",
    "Assam":"AS","Bihar":"BR","Chandigarh":"CH","Chhattisgarh":"CG",
    "Dadra and Nagar Haveli":"DN","Daman and Diu":"DD","Delhi":"DL","Goa":"GA",
    "Gujarat":"GJ","Haryana":"HR","Himachal Pradesh":"HP","Jammu and Kashmir":"JK",
    "Jharkhand":"JH","Karnataka":"KA","Kerala":"KL","Ladakh":"LA","Lakshadweep":"LD",
    "Madhya Pradesh":"MP","Maharashtra":"MH","Manipur":"MN","Meghalaya":"ML",
    "Mizoram":"MZ","Nagaland":"NL","Odisha":"OD","Puducherry":"PY","Punjab":"PB",
    "Rajasthan":"RJ","Sikkim":"SK","Tamil Nadu":"TN","Telangana":"TS","Tripura":"TR",
    "Uttar Pradesh":"UP","Uttarakhand":"UK","West Bengal":"WB",
}

def fetch(url):
    for attempt in range(3):
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            return r.text
        except Exception as e:
            print(f"    warning attempt {attempt+1}: {e}")
            time.sleep(3)
    return None

def get_all_districts():
    url = f"{BASE}/education/icse-schools.html"
    print(f"Fetching master district list...")
    html = fetch(url)
    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")
    results = []
    current_state = None
    for tag in soup.find_all(True):
        text = tag.get_text(strip=True)
        state_m = re.match(r"ICSE Schools in (.+?)\s*\(\d+\)", text, re.IGNORECASE)
        if state_m and tag.name in ("h2","h3","h4","b","strong","span","div","p","li"):
            candidate = state_m.group(1).strip()
            if candidate in STATE_CODE:
                current_state = candidate
            continue
        if tag.name == "a" and current_state:
            href = tag.get("href", "")
            slug_m = re.search(r"icse-schools-in-(.+?)\.html", href)
            if slug_m:
                d_slug = slug_m.group(1)
                d_name = text.replace("ICSE Schools in ","").strip()
                if d_slug and d_name:
                    results.append((current_state, d_name, d_slug))
    print(f"Found {len(results)} districts across {len(set(r[0] for r in results))} states.\n")
    return results

def parse_schools(html, district_name, state_name):
    soup = BeautifulSoup(html, "html.parser")
    schools = []
    for li in soup.find_all("li"):
        a = li.find("a", href=re.compile(r"/education/.+-s\d+\.html"))
        if not a:
            continue
        school_name = a.get_text(strip=True)
        if not school_name:
            continue
        full_text = li.get_text(" ", strip=True)
        details = full_text.replace(school_name, "", 1).strip().lstrip(".")
        phone_m = re.search(r"([\d][0-9\-\+\s,]{7,})$", details)
        phone   = phone_m.group(1).strip() if phone_m else ""
        address = details[:phone_m.start()].strip() if phone_m else details
        address = re.sub(r"\s*\.\.\.$", "", address).strip()
        schools.append({"school_name": school_name, "district": district_name, "state": state_name, "address": address})
    return schools

def scrape_district(state_name, district_name, district_slug, state_code, seq_counters):
    all_schools = []
    page = 1
    state_slug = state_name.lower().replace(" ", "-").replace("&", "and")
    while True:
        url = (f"{BASE}/education/icse-schools-in-{district_slug}.html" if page == 1
               else f"{BASE}/education/search.php?p=1&mode=school&state={state_slug}&district={district_slug}&school_type=icse&page={page}")
        html = fetch(url)
        if not html:
            break
        schools = parse_schools(html, district_name, state_name)
        if not schools:
            break
        for school in schools:
            key = f"ICS-{state_code}"
            seq_counters[key] = seq_counters.get(key, 0) + 1
            school.update({"cb_code": f"CB-ICS-{state_code}-{seq_counters[key]:05d}","board":"CISCE",
                           "affiliation_id":"","level":"ICSE/ISC","year_founded":"","pincode":"","email":"","management":""})
        all_schools.extend(schools)
        soup = BeautifulSoup(html, "html.parser")
        if soup.find("a", string=str(page + 1)):
            page += 1
            time.sleep(REQUEST_DELAY)
        else:
            break
    return all_schools

def scrape_all_cisce():
    print("\n=== SCRAPING CISCE SCHOOLS (prokerala.com) v2 ===\n")
    districts = get_all_districts()
    if not districts:
        return []
    all_schools = []
    seq_counters = {}
    current_state = None
    for state_name, district_name, district_slug in districts:
        if state_name != current_state:
            if current_state:
                total = sum(1 for s in all_schools if s["state"] == current_state)
                print(f"  Total {current_state}: {total}\n")
            current_state = state_name
            print(f">> {state_name}")
        state_code = STATE_CODE.get(state_name, "XX")
        schools = scrape_district(state_name, district_name, district_slug, state_code, seq_counters)
        print(f"  {district_name}: {len(schools)} schools")
        all_schools.extend(schools)
        time.sleep(REQUEST_DELAY)
    if current_state:
        total = sum(1 for s in all_schools if s["state"] == current_state)
        print(f"  Total {current_state}: {total}\n")
    seen = set()
    deduped = []
    for s in all_schools:
        key = (s["school_name"].lower().strip(), s["district"].lower().strip())
        if key not in seen:
            seen.add(key)
            deduped.append(s)
    print(f"\nCISCE total after dedup: {len(deduped)} schools")
    return deduped

COLUMNS = [("CB Code","cb_code",22),("Board","board",10),("School Name","school_name",45),
           ("District","district",20),("State","state",20),("Affiliation ID","affiliation_id",15),
           ("Level","level",20),("Year Founded","year_founded",14),("Address","address",45),
           ("PIN Code","pincode",12),("Email","email",35),("Management","management",45)]
TEAL="0B5C45"; AMBER="E8922A"; WHITE="FFFFFF"; LGREY="F5F5F5"; MGREY="E0E0E0"

def save_csv(schools, filepath, append=False):
    fieldnames = [f for _,f,_ in COLUMNS]
    file_exists = os.path.exists(filepath)
    mode = "a" if (append and file_exists) else "w"
    with open(filepath, mode, newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if not (append and file_exists):
            w.writeheader()
        w.writerows(schools)
    print(f"  Saved: {filepath}")

def rebuild_excel(csv_path, xlsx_path):
    print(f"\nRebuilding Excel...")
    with open(csv_path, newline="", encoding="utf-8") as f:
        all_schools = list(csv.DictReader(f))
    print(f"  Total rows: {len(all_schools)}")
    wb = Workbook()
    ws = wb.active
    ws.title = "School Directory"
    ws.freeze_panes = "A2"
    hf = Font(name="Arial", bold=True, color=WHITE, size=10)
    hfill = PatternFill("solid", start_color=TEAL, end_color=TEAL)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr = Border(bottom=Side(style="thin", color=MGREY), right=Side(style="thin", color=MGREY))
    for ci,(col_name,_,cw) in enumerate(COLUMNS,1):
        c = ws.cell(row=1, column=ci, value=col_name)
        c.font, c.fill, c.alignment = hf, hfill, ha
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[1].height = 24
    even = PatternFill("solid", start_color=LGREY, end_color=LGREY)
    df = Font(name="Arial", size=9)
    for ri, school in enumerate(all_schools, 2):
        for ci,(_,field,_) in enumerate(COLUMNS,1):
            c = ws.cell(row=ri, column=ci, value=school.get(field,""))
            c.font = df; c.border = bdr; c.alignment = Alignment(vertical="center")
            if ri % 2 == 0: c.fill = even
    ws2 = wb.create_sheet("Summary by State")
    ws2.freeze_panes = "A2"
    for ci,h in enumerate(["State","CBSE","CISCE","Total"],1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha
        ws2.column_dimensions[get_column_letter(ci)].width = 25
    counts = {}
    for s in all_schools:
        st, bd = s.get("state",""), s.get("board","")
        if st not in counts: counts[st] = {"CBSE":0,"CISCE":0}
        if bd in counts[st]: counts[st][bd] += 1
    for ri,(st,c) in enumerate(sorted(counts.items()),2):
        ws2.cell(row=ri,column=1,value=st).font=df
        ws2.cell(row=ri,column=2,value=c["CBSE"]).font=df
        ws2.cell(row=ri,column=3,value=c["CISCE"]).font=df
        ws2.cell(row=ri,column=4,value=c["CBSE"]+c["CISCE"]).font=Font(name="Arial",size=9,bold=True)
    tr = len(counts)+2
    ct = sum(v["CBSE"] for v in counts.values())
    ci2 = sum(v["CISCE"] for v in counts.values())
    afill = PatternFill("solid", start_color=AMBER, end_color=AMBER)
    bw = Font(name="Arial", size=9, bold=True, color=WHITE)
    for ci,val in enumerate(["TOTAL",ct,ci2,ct+ci2],1):
        c = ws2.cell(row=tr, column=ci, value=val)
        c.font=bw; c.fill=afill; c.alignment=Alignment(horizontal="center")
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

def main():
    cisce = scrape_all_cisce()
    if not cisce:
        print("No schools scraped.")
        return
    save_csv(cisce, "catalystbox_cisce_schools.csv")
    save_csv(cisce, "catalystbox_school_directory.csv", append=True)
    rebuild_excel("catalystbox_school_directory.csv", "catalystbox_school_directory.xlsx")
    print(f"\nDone! {len(cisce)} CISCE schools added.")

if __name__ == "__main__":
    main()
