"""
CatalystBox School Directory Scraper
=====================================
Scrapes CBSE schools from cbseschool.org (mirrors cbseaff.nic.in)
and CISCE schools from cisce.org.

Usage:
    pip install requests beautifulsoup4 openpyxl
    python catalystbox_school_scraper.py

Output:
    catalystbox_school_directory.xlsx  — full directory with CB codes
    catalystbox_school_directory.csv   — CSV backup
"""

import requests
import re
import csv
import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── CONFIG ────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
}

REQUEST_DELAY = 1.2   # seconds between requests — be polite
TIMEOUT       = 20    # seconds per request

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


# ─── STATE METADATA ────────────────────────────────────────────────────────────

CBSE_STATES = [
    ("andaman-nicobar",       "Andaman & Nicobar",   "AN"),
    ("andhra-pradesh",        "Andhra Pradesh",       "AP"),
    ("arunachal-pradesh",     "Arunachal Pradesh",    "AR"),
    ("assam",                 "Assam",                "AS"),
    ("bihar",                 "Bihar",                "BR"),
    ("chandigarh",            "Chandigarh",           "CH"),
    ("chattisgarh",           "Chattisgarh",          "CG"),
    ("dadar-nagar-haveli",    "Dadra & Nagar Haveli", "DN"),
    ("daman-diu",             "Daman & Diu",          "DD"),
    ("delhi",                 "Delhi",                "DL"),
    ("goa",                   "Goa",                  "GA"),
    ("gujarat",               "Gujarat",              "GJ"),
    ("haryana",               "Haryana",              "HR"),
    ("himachal-pradesh",      "Himachal Pradesh",     "HP"),
    ("jammu-kashmir",         "Jammu & Kashmir",      "JK"),
    ("jharkhand",             "Jharkhand",            "JH"),
    ("karnataka",             "Karnataka",            "KA"),
    ("kerala",                "Kerala",               "KL"),
    ("ladakh",                "Ladakh",               "LA"),
    ("lakshadweep",           "Lakshadweep",          "LD"),
    ("madhya-pradesh",        "Madhya Pradesh",       "MP"),
    ("maharashtra",           "Maharashtra",          "MH"),
    ("manipur",               "Manipur",              "MN"),
    ("meghalaya",             "Meghalaya",            "ML"),
    ("mizoram",               "Mizoram",              "MZ"),
    ("nagaland",              "Nagaland",             "NL"),
    ("odisha",                "Odisha",               "OD"),
    ("puducherry",            "Puducherry",           "PY"),
    ("punjab",                "Punjab",               "PB"),
    ("rajasthan",             "Rajasthan",            "RJ"),
    ("sikkim",                "Sikkim",               "SK"),
    ("tamil-nadu",            "Tamil Nadu",           "TN"),
    ("telangana",             "Telangana",            "TS"),
    ("tripura",               "Tripura",              "TR"),
    ("uttar-pradesh",         "Uttar Pradesh",        "UP"),
    ("uttarakhand",           "Uttarakhand",          "UK"),
    ("west-bengal",           "West Bengal",          "WB"),
]

STATE_CODE_MAP = {slug: code for slug, _, code in CBSE_STATES}
STATE_NAME_MAP = {slug: name for slug, name, _ in CBSE_STATES}


# ─── CBSE SCRAPER ──────────────────────────────────────────────────────────────

def fetch_page(url):
    """Fetch a URL with retry logic."""
    for attempt in range(3):
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            return r.text
        except Exception as e:
            print(f"    ⚠  Attempt {attempt+1} failed for {url}: {e}")
            time.sleep(3)
    return None


def parse_cbse_page(html, state_name):
    """Parse all school entries from a cbseschool.org state/page HTML."""
    soup = BeautifulSoup(html, "html.parser")
    schools = []

    # Each school is an <h2> followed by a <p> with details
    # Pattern: Founded in YEAR, NAME is LEVEL, affiliated to CBSE.
    #          Affiliation ID is XXXXX. Address of the school is: ADDR.
    #          PIN Code: YYYYY. Email address of the school is EMAIL.
    #          The school is being managed by TRUST.
    for h2 in soup.find_all("h2"):
        text_parts = []
        sibling = h2.find_next_sibling()
        while sibling and sibling.name not in ("h2", "h3"):
            if sibling.name == "p":
                text_parts.append(sibling.get_text(" ", strip=True))
            sibling = sibling.find_next_sibling()

        full_text = " ".join(text_parts)
        school_name_raw = h2.get_text(strip=True)

        # Extract district from school name header "School Name - District"
        district = ""
        if " - " in school_name_raw:
            parts = school_name_raw.rsplit(" - ", 1)
            school_name = parts[0].strip()
            district = parts[1].strip()
        else:
            school_name = school_name_raw.strip()

        # Extract fields from paragraph text
        affiliation_id = _extract(r"Affiliation ID is\s*([\d]+)", full_text)
        year_founded   = _extract(r"Founded in\s*(\d{4})", full_text)
        level          = _extract(r"is a\s*([\w\s]+?),\s*affiliated to CBSE", full_text)
        address        = _extract(r"Address of the school is:\s*(.+?)(?:\.\s*PIN|$)", full_text)
        pincode        = _extract(r"PIN Code:\s*(\d{6})", full_text)
        email          = _extract(r"Email address of the school is\s*([\S]+@[\S]+)", full_text)
        management     = _extract(r"managed by\s*(.+?)(?:\.\s*Read More|$)", full_text)

        if not school_name or school_name.lower() in ("home", "about us"):
            continue

        schools.append({
            "school_name":     school_name,
            "district":        district,
            "state":           state_name,
            "affiliation_id":  affiliation_id,
            "level":           level.strip() if level else "",
            "year_founded":    year_founded,
            "address":         address.strip() if address else "",
            "pincode":         pincode,
            "email":           email.lower().strip() if email else "",
            "management":      management.strip() if management else "",
            "board":           "CBSE",
        })

    return schools


def _extract(pattern, text):
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else ""


def get_next_page_url(html, current_url):
    """Find the 'Next' pagination link if present."""
    soup = BeautifulSoup(html, "html.parser")
    next_link = soup.find("a", class_="next")
    if next_link and next_link.get("href"):
        return next_link["href"]
    # Also look for numbered pagination
    pagination = soup.find("div", class_=re.compile(r"pagination|nav-links", re.I))
    if pagination:
        links = pagination.find_all("a")
        for i, link in enumerate(links):
            if link.get("href") and link.get("href") == current_url:
                if i + 1 < len(links):
                    return links[i + 1].get("href")
    return None


def scrape_cbse_state(slug, state_name, seq_counters):
    """Scrape all pages of a state and return school records with CB codes."""
    # cbseschool.org uses WordPress pagination:
    # Page 1: /location/uttar-pradesh/
    # Page 2: /location/uttar-pradesh/page/2/
    # Page N: /location/uttar-pradesh/page/N/
    state_code = STATE_CODE_MAP.get(slug, "XX")
    all_schools = []
    page = 1

    while True:
        if page == 1:
            url = f"https://www.cbseschool.org/location/{slug}/"
        else:
            url = f"https://www.cbseschool.org/location/{slug}/page/{page}/"

        print(f"  Fetching page {page}: {url}")
        html = fetch_page(url)
        if not html:
            print(f"  ✗ Could not fetch {url} — stopping state")
            break

        # WordPress returns 404 content or redirects when page exceeds total
        # Detect this: if the page redirects back to page 1 or returns no schools
        soup = BeautifulSoup(html, "html.parser")

        # Stop if WordPress signals 404 / "page not found"
        title = soup.find("title")
        if title and ("404" in title.text or "Page not found" in title.text.lower()):
            print(f"    Reached end at page {page} (404)")
            break

        schools = parse_cbse_page(html, state_name)
        print(f"    Found {len(schools)} schools")

        # If 0 schools found on page > 1, we've gone past the end
        if len(schools) == 0:
            if page > 1:
                print(f"    No schools found — end of state listing")
            break

        for school in schools:
            key = f"CBS-{state_code}"
            seq_counters[key] = seq_counters.get(key, 0) + 1
            school["cb_code"] = f"CB-CBS-{state_code}-{seq_counters[key]:05d}"

        all_schools.extend(schools)
        page += 1
        time.sleep(REQUEST_DELAY)

    return all_schools


def scrape_all_cbse():
    """Scrape all CBSE states."""
    print("\n╔══════════════════════════════════════════╗")
    print("║  SCRAPING CBSE SCHOOLS (cbseschool.org)  ║")
    print("╚══════════════════════════════════════════╝\n")

    all_schools = []
    seq_counters = {}

    for slug, state_name, state_code in CBSE_STATES:
        print(f"\n▶ {state_name} ({state_code})")
        try:
            schools = scrape_cbse_state(slug, state_name, seq_counters)
            print(f"  ✓ Total: {len(schools)} schools in {state_name}")
            all_schools.extend(schools)
            time.sleep(REQUEST_DELAY)
        except Exception as e:
            print(f"  ✗ Error scraping {state_name}: {e}")

    print(f"\n✓ CBSE total: {len(all_schools)} schools\n")
    return all_schools


# ─── CISCE SCRAPER ─────────────────────────────────────────────────────────────

CISCE_STATE_CODES = {
    "Andhra Pradesh": "AP", "Arunachal Pradesh": "AR", "Assam": "AS",
    "Bihar": "BR", "Chhattisgarh": "CG", "Goa": "GA", "Gujarat": "GJ",
    "Haryana": "HR", "Himachal Pradesh": "HP", "Jammu and Kashmir": "JK",
    "Jharkhand": "JH", "Karnataka": "KA", "Kerala": "KL",
    "Madhya Pradesh": "MP", "Maharashtra": "MH", "Manipur": "MN",
    "Meghalaya": "ML", "Mizoram": "MZ", "Nagaland": "NL", "Odisha": "OD",
    "Punjab": "PB", "Rajasthan": "RJ", "Sikkim": "SK", "Tamil Nadu": "TN",
    "Telangana": "TS", "Tripura": "TR", "Uttar Pradesh": "UP",
    "Uttarakhand": "UK", "West Bengal": "WB", "Delhi": "DL",
    "Chandigarh": "CH", "Puducherry": "PY",
}


def scrape_cisce():
    """Scrape CISCE schools from cisce.org/find-a-school."""
    print("╔══════════════════════════════════════════╗")
    print("║  SCRAPING CISCE SCHOOLS (cisce.org)      ║")
    print("╚══════════════════════════════════════════╝\n")

    schools = []
    seq_counters = {}
    base = "https://cisce.org"
    url  = "https://cisce.org/find-a-school"

    try:
        html = fetch_page(url)
        if not html:
            print("  ✗ Could not reach cisce.org/find-a-school")
            return []

        soup = BeautifulSoup(html, "html.parser")

        # CISCE uses a search form — try fetching JSON API or table results
        # The site has a state dropdown; we enumerate states
        state_select = soup.find("select", {"name": re.compile(r"state", re.I)})
        state_options = []
        if state_select:
            state_options = [
                (opt.get("value"), opt.get_text(strip=True))
                for opt in state_select.find_all("option")
                if opt.get("value") and opt.get("value") != "0"
            ]
            print(f"  Found {len(state_options)} states in CISCE dropdown")
        else:
            # Fallback: try direct school listing pages
            print("  Trying direct CISCE school listing...")
            school_links = soup.find_all("a", href=re.compile(r"/school/", re.I))
            print(f"  Found {len(school_links)} direct school links")

        # Try form POST to get schools per state
        form = soup.find("form")
        form_action = base + (form.get("action", "") if form else "")

        for state_val, state_name in state_options:
            print(f"  ▶ {state_name}")
            try:
                resp = SESSION.post(
                    form_action,
                    data={"state": state_val, "school_name": "", "submit": "Search"},
                    timeout=TIMEOUT,
                )
                state_soup = BeautifulSoup(resp.text, "html.parser")
                rows = state_soup.find_all("tr")
                state_code = CISCE_STATE_CODES.get(state_name, "XX")
                found = 0

                for row in rows:
                    cols = row.find_all("td")
                    if len(cols) >= 3:
                        name   = cols[0].get_text(strip=True)
                        school_code_raw = cols[1].get_text(strip=True) if len(cols) > 1 else ""
                        city   = cols[2].get_text(strip=True) if len(cols) > 2 else ""

                        if not name or name.lower() in ("school name", "name"):
                            continue

                        key = f"ICS-{state_code}"
                        seq_counters[key] = seq_counters.get(key, 0) + 1

                        schools.append({
                            "cb_code":        f"CB-ICS-{state_code}-{seq_counters[key]:05d}",
                            "school_name":    name,
                            "district":       city,
                            "state":          state_name,
                            "affiliation_id": school_code_raw,
                            "level":          "ICSE/ISC",
                            "year_founded":   "",
                            "address":        "",
                            "pincode":        "",
                            "email":          "",
                            "management":     "",
                            "board":          "CISCE",
                        })
                        found += 1

                print(f"    ✓ {found} schools")
                time.sleep(REQUEST_DELAY)

            except Exception as e:
                print(f"    ✗ Error: {e}")

    except Exception as e:
        print(f"  ✗ CISCE scrape failed: {e}")

    print(f"\n✓ CISCE total: {len(schools)} schools\n")
    return schools


# ─── EXCEL EXPORT ──────────────────────────────────────────────────────────────

COLUMNS = [
    ("CB Code",        "cb_code",        22),
    ("Board",          "board",          10),
    ("School Name",    "school_name",    45),
    ("District",       "district",       20),
    ("State",          "state",          20),
    ("Affiliation ID", "affiliation_id", 15),
    ("Level",          "level",          20),
    ("Year Founded",   "year_founded",   14),
    ("Address",        "address",        45),
    ("PIN Code",       "pincode",        12),
    ("Email",          "email",          35),
    ("Management",     "management",     45),
]

TEAL   = "0B5C45"
AMBER  = "E8922A"
WHITE  = "FFFFFF"
LGREY  = "F5F5F5"
MGREY  = "E0E0E0"


def export_excel(schools, filepath):
    print(f"\n📊 Exporting {len(schools)} schools to Excel...")
    wb = Workbook()

    # ── Sheet 1: Full Directory ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "School Directory"
    ws.freeze_panes = "A2"

    header_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    header_fill  = PatternFill("solid", start_color=TEAL, end_color=TEAL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin  = Border(
        bottom=Side(style="thin", color=MGREY),
        right=Side(style="thin", color=MGREY),
    )

    # Header row
    for col_idx, (col_name, _, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 24

    # Data rows
    even_fill = PatternFill("solid", start_color=LGREY, end_color=LGREY)
    data_font = Font(name="Arial", size=9)

    for row_idx, school in enumerate(schools, start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=school.get(field, ""))
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # ── Sheet 2: Summary by State ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary by State")
    ws2.freeze_panes = "A2"

    summary_headers = ["State", "CBSE", "CISCE", "Total"]
    for col_idx, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws2.column_dimensions[get_column_letter(col_idx)].width = 25

    # Build state summary
    state_counts = {}
    for school in schools:
        state = school["state"]
        board = school["board"]
        if state not in state_counts:
            state_counts[state] = {"CBSE": 0, "CISCE": 0}
        if board in state_counts[state]:
            state_counts[state][board] += 1

    for row_idx, (state, counts) in enumerate(sorted(state_counts.items()), start=2):
        cbse_count  = counts.get("CBSE", 0)
        cisce_count = counts.get("CISCE", 0)
        total       = cbse_count + cisce_count
        ws2.cell(row=row_idx, column=1, value=state).font = data_font
        ws2.cell(row=row_idx, column=2, value=cbse_count).font = data_font
        ws2.cell(row=row_idx, column=3, value=cisce_count).font = data_font
        ws2.cell(row=row_idx, column=4, value=total).font = Font(name="Arial", size=9, bold=True)

    # Totals row
    total_rows = len(state_counts) + 2
    cbse_total  = sum(c["CBSE"]  for c in state_counts.values())
    cisce_total = sum(c["CISCE"] for c in state_counts.values())
    amber_fill  = PatternFill("solid", start_color=AMBER, end_color=AMBER)
    bold_font   = Font(name="Arial", size=9, bold=True, color=WHITE)

    for col_idx, val in enumerate([
        "TOTAL", cbse_total, cisce_total, cbse_total + cisce_total
    ], start=1):
        cell = ws2.cell(row=total_rows, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = amber_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Sheet 3: CB Code Legend ────────────────────────────────────────────────
    ws3 = wb.create_sheet("CB Code Format")
    ws3["A1"] = "CatalystBox School Code System"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color=TEAL)
    ws3["A3"] = "Format:  CB - [BOARD] - [STATE] - [SEQUENCE]"
    ws3["A3"].font = Font(name="Arial", size=11, bold=True)
    ws3["A5"] = "Board Codes"
    ws3["A5"].font = Font(name="Arial", bold=True, size=10)

    board_codes = [
        ("CBS", "CBSE"),
        ("ICS", "CISCE (ICSE/ISC)"),
        ("UPB", "UP Board"),
        ("SSC", "Maharashtra SSC"),
        ("TBT", "Tamil Nadu Board"),
        ("KAB", "Karnataka Board"),
        ("RAJ", "Rajasthan Board"),
        ("GEN", "Other State Boards"),
    ]
    for row_offset, (code, name) in enumerate(board_codes, start=6):
        ws3.cell(row=row_offset, column=1, value=code).font = Font(
            name="Courier New", bold=True, size=10, color=AMBER
        )
        ws3.cell(row=row_offset, column=2, value=name).font = Font(name="Arial", size=10)

    ws3["A15"] = "State Codes"
    ws3["A15"].font = Font(name="Arial", bold=True, size=10)
    for row_offset, (_, state_name, state_code) in enumerate(CBSE_STATES, start=16):
        ws3.cell(row=row_offset, column=1, value=state_code).font = Font(
            name="Courier New", bold=True, size=9, color=TEAL
        )
        ws3.cell(row=row_offset, column=2, value=state_name).font = Font(name="Arial", size=9)

    ws3["A55"] = "Example:  CB-CBS-UP-00042 = CBSE school, Uttar Pradesh, #42 in UP"
    ws3["A55"].font = Font(name="Arial", size=10, italic=True)

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 35

    wb.save(filepath)
    print(f"  ✓ Saved: {filepath}")


def export_csv(schools, filepath):
    print(f"📄 Exporting CSV backup: {filepath}")
    fieldnames = [field for _, field, _ in COLUMNS]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(schools)
    print(f"  ✓ Saved: {filepath}")


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "═"*50)
    print("  CatalystBox School Directory Builder")
    print("═"*50)

    all_schools = []

    # CBSE
    cbse_schools = scrape_all_cbse()
    all_schools.extend(cbse_schools)

    # CISCE
    cisce_schools = scrape_cisce()
    all_schools.extend(cisce_schools)

    if not all_schools:
        print("\n✗ No schools scraped. Check your internet connection.")
        return

    print(f"\n{'═'*50}")
    print(f"  TOTAL SCHOOLS: {len(all_schools)}")
    print(f"  CBSE:  {sum(1 for s in all_schools if s['board'] == 'CBSE')}")
    print(f"  CISCE: {sum(1 for s in all_schools if s['board'] == 'CISCE')}")
    print(f"{'═'*50}\n")

    xlsx_path = "catalystbox_school_directory.xlsx"
    csv_path  = "catalystbox_school_directory.csv"

    export_excel(all_schools, xlsx_path)
    export_csv(all_schools, csv_path)

    print("\n✅ Done! Files saved to current directory.")
    print(f"   {xlsx_path}")
    print(f"   {csv_path}")


if __name__ == "__main__":
    main()
